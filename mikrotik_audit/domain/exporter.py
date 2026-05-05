from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import AppConfig
from constants.error_codes import FirmwareErrorCode
from constants.statuses import AuditStatus
from models import AuditResult


class ExcelExporter:
    INVENTORY_HEADERS = [
        "ip",
        "identity",
        "status",
        "inventory_status",
        "inventory_severity",
        "phpipam_match_type",
        "phpipam_hostname",
        "phpipam_ip",
        "version",
        "board_name",
        "platform",
        "architecture",
        "ospf_instance_count",
        "ospf_neighbor_count",
        "bridge_count",
        "bridge_port_count",
        "scheduler_count",
        "dhcp_server_count",
        "dhcp_client_count",
        "ssh_port_value",
        "winbox_port_value",
        "firewall_filter_count",
        "firewall_nat_count",
        "route_count",
        "default_route_count",
        "vlan_count",
        "radius_count",
        "watchdog_enabled",
        "ping",
        "ssh_port",
        "auth_method",
        "firmware_target_version",
        "firmware_error",
        "phpipam_note",
    ]
    TOPOLOGY_HEADERS = [
        "ip",
        "identity",
        "board_name",
        "mac_address",
        "uplink_interface",
        "uplink_mac",
        "neighbor_identity",
        "neighbor_address",
        "neighbor_interface",
        "neighbor_mac",
    ]
    PHPIPAM_MISMATCH_HEADERS = [
        "ip",
        "identity",
        "inventory_status",
        "inventory_severity",
        "phpipam_match_type",
        "phpipam_hostname",
        "phpipam_ip",
        "phpipam_address_id",
        "phpipam_description",
        "phpipam_note",
        "status",
        "version",
        "board_name",
    ]
    RAW_HEADERS = AuditResult.EXPORT_HEADERS
    HEADER_FILL = PatternFill("solid", fgColor="1F1F1F")
    SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
    STATUS_FILLS = {
        AuditStatus.SSH_OK.value: PatternFill("solid", fgColor="C6EFCE"),
        AuditStatus.FALLBACK_OK.value: PatternFill("solid", fgColor="FFF2CC"),
        AuditStatus.AUTH_FAILED.value: PatternFill("solid", fgColor="F4CCCC"),
        AuditStatus.OFFLINE.value: PatternFill("solid", fgColor="D9D9D9"),
        AuditStatus.SSH_CLOSED.value: PatternFill("solid", fgColor="FCE5CD"),
    }
    SEVERITY_FILLS = {
        "INFO": PatternFill("solid", fgColor="E2F0D9"),
        "WARNING": PatternFill("solid", fgColor="FFF2CC"),
        "ERROR": PatternFill("solid", fgColor="F4CCCC"),
    }
    ISSUE_HEADERS = [
        "ip",
        "identity",
        "status",
        "inventory_status",
        "inventory_severity",
        "phpipam_match_type",
        "phpipam_hostname",
        "phpipam_ip",
        "version",
        "board_name",
        "firmware_error",
        "phpipam_note",
    ]

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    @staticmethod
    def autosize_worksheet(ws: Worksheet) -> None:
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)

            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    @classmethod
    def _resolve_status_fill(cls, status: str) -> PatternFill | None:
        if status.startswith(AuditStatus.SSH_OK.value):
            return cls.STATUS_FILLS[AuditStatus.SSH_OK.value]
        if status.startswith(AuditStatus.FALLBACK_OK.value):
            return cls.STATUS_FILLS[AuditStatus.FALLBACK_OK.value]
        return cls.STATUS_FILLS.get(status)

    @classmethod
    def apply_styles(cls, ws: Worksheet, row_idx: int, status: str) -> None:
        fill = cls._resolve_status_fill(status)
        if fill:
            for cell in ws[row_idx]:
                cell.fill = fill

    @classmethod
    def _style_header_row(cls, ws: Worksheet, row_idx: int) -> None:
        for cell in ws[row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = cls.HEADER_FILL

    @classmethod
    def _style_section_row(cls, ws: Worksheet, row_idx: int) -> None:
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.fill = cls.SECTION_FILL

    @classmethod
    def _apply_issue_styles(cls, ws: Worksheet, row_idx: int, issue: dict[str, Any]) -> None:
        severity = str(issue.get("inventory_severity", "")).upper()
        fill = cls.SEVERITY_FILLS.get(severity)

        if fill is None:
            fill = cls._resolve_status_fill(str(issue.get("status", "")))

        if fill is None:
            return

        for cell in ws[row_idx]:
            cell.fill = fill

    def _build_summary_rows(self, results: list[AuditResult]) -> list[tuple[str, Any]]:
        status_counts = Counter()
        flag_counts = Counter()
        firmware_error_counts = Counter()
        inventory_status_counts = Counter()
        inventory_severity_counts = Counter()
        inventory_match_counts = Counter()

        for result in results:
            if result.ping:
                flag_counts["alive"] += 1
            if result.radius_added:
                flag_counts["radius_added"] += 1
            if result.radius_recreated:
                flag_counts["radius_recreated"] += 1
            if result.radius_present_after:
                flag_counts["radius_present_after"] += 1
            if result.aaa_enabled:
                flag_counts["aaa_enabled"] += 1
            if result.aaa_present_after:
                flag_counts["aaa_present_after"] += 1
            if result.firmware_upload_needed:
                flag_counts["firmware_upload_needed"] += 1
            if result.firmware_uploaded:
                flag_counts["firmware_uploaded"] += 1
            if result.firmware_already_present:
                flag_counts["firmware_already_present"] += 1
            if result.firmware_reboot_sent:
                flag_counts["firmware_reboot_sent"] += 1

            if result.status.startswith(AuditStatus.SSH_OK.value):
                status_counts["ssh_ok"] += 1
            elif result.status.startswith(AuditStatus.FALLBACK_OK.value):
                status_counts["fallback_ok"] += 1
            elif result.status == AuditStatus.AUTH_FAILED.value:
                status_counts["auth_failed"] += 1
            elif result.status == AuditStatus.SSH_CLOSED.value:
                status_counts["ssh_closed"] += 1
            elif result.status == AuditStatus.OFFLINE.value:
                status_counts["offline"] += 1

            if result.firmware_error:
                firmware_error_counts[result.firmware_error] += 1
            if result.inventory_status:
                inventory_status_counts[result.inventory_status] += 1
            if result.inventory_severity:
                inventory_severity_counts[result.inventory_severity] += 1
            if result.phpipam_match_type:
                inventory_match_counts[result.phpipam_match_type] += 1

        total_hosts = len(results)
        alive = flag_counts["alive"]
        ok_inventory = inventory_status_counts["OK"]

        return [
            ("total_hosts", total_hosts),
            ("alive", alive),
            (
                "alive_percent",
                round((alive / total_hosts) * 100, 2) if total_hosts else 0,
            ),
            ("ssh_ok", status_counts["ssh_ok"]),
            ("fallback_ok", status_counts["fallback_ok"]),
            ("auth_failed", status_counts["auth_failed"]),
            ("ssh_closed", status_counts["ssh_closed"]),
            ("offline", status_counts["offline"]),
            ("radius_added", flag_counts["radius_added"]),
            ("radius_recreated", flag_counts["radius_recreated"]),
            ("radius_present_after", flag_counts["radius_present_after"]),
            ("aaa_enabled", flag_counts["aaa_enabled"]),
            ("aaa_present_after", flag_counts["aaa_present_after"]),
            ("firmware_upload_needed", flag_counts["firmware_upload_needed"]),
            ("firmware_uploaded", flag_counts["firmware_uploaded"]),
            ("firmware_already_present", flag_counts["firmware_already_present"]),
            ("firmware_reboot_sent", flag_counts["firmware_reboot_sent"]),
            (
                "firmware_same_version",
                firmware_error_counts[FirmwareErrorCode.SAME_VERSION.value],
            ),
            (
                "firmware_upload_failed",
                firmware_error_counts[FirmwareErrorCode.UPLOAD_FAILED.value],
            ),
            (
                "firmware_local_not_found",
                firmware_error_counts[FirmwareErrorCode.LOCAL_FIRMWARE_NOT_FOUND.value],
            ),
            (
                "inventory_ok",
                ok_inventory,
            ),
            (
                "inventory_compliance_percent",
                round((ok_inventory / total_hosts) * 100, 2) if total_hosts else 0,
            ),
            (
                "inventory_hostname_mismatch",
                inventory_status_counts["HOSTNAME_MISMATCH"],
            ),
            (
                "inventory_hostname_partial_match",
                inventory_status_counts["HOSTNAME_PARTIAL_MATCH"],
            ),
            (
                "inventory_hostname_incomplete",
                inventory_status_counts["HOSTNAME_INCOMPLETE"],
            ),
            (
                "inventory_not_found",
                inventory_status_counts["NOT_FOUND"],
            ),
            (
                "inventory_duplicate",
                inventory_status_counts["DUPLICATE"],
            ),
            (
                "severity_info",
                inventory_severity_counts["INFO"],
            ),
            (
                "severity_warning",
                inventory_severity_counts["WARNING"],
            ),
            (
                "severity_error",
                inventory_severity_counts["ERROR"],
            ),
            ("match_ip", inventory_match_counts["ip"]),
            ("match_hostname", inventory_match_counts["hostname"]),
            ("match_partial_hostname", inventory_match_counts["partial_hostname"]),
            ("match_not_found", inventory_match_counts["not_found"]),
        ]

    def _build_report_sections(self, results: list[AuditResult]) -> list[tuple[str, list[tuple[str, Any]]]]:
        summary_map = dict(self._build_summary_rows(results))

        return [
            (
                "report",
                [
                    ("generated_at_utc", datetime.now(timezone.utc).isoformat()),
                    ("inventory_file", self.config.inventory_file),
                    ("output_xlsx", self.config.output_xlsx),
                    ("total_hosts", summary_map["total_hosts"]),
                    ("alive_percent", summary_map["alive_percent"]),
                    ("inventory_compliance_percent", summary_map["inventory_compliance_percent"]),
                ],
            ),
            (
                "audit",
                [
                    ("alive", summary_map["alive"]),
                    ("ssh_ok", summary_map["ssh_ok"]),
                    ("fallback_ok", summary_map["fallback_ok"]),
                    ("auth_failed", summary_map["auth_failed"]),
                    ("ssh_closed", summary_map["ssh_closed"]),
                    ("offline", summary_map["offline"]),
                ],
            ),
            (
                "remediation",
                [
                    ("radius_added", summary_map["radius_added"]),
                    ("radius_recreated", summary_map["radius_recreated"]),
                    ("radius_present_after", summary_map["radius_present_after"]),
                    ("aaa_enabled", summary_map["aaa_enabled"]),
                    ("aaa_present_after", summary_map["aaa_present_after"]),
                    ("firmware_upload_needed", summary_map["firmware_upload_needed"]),
                    ("firmware_uploaded", summary_map["firmware_uploaded"]),
                    ("firmware_already_present", summary_map["firmware_already_present"]),
                    ("firmware_reboot_sent", summary_map["firmware_reboot_sent"]),
                    ("firmware_same_version", summary_map["firmware_same_version"]),
                    ("firmware_upload_failed", summary_map["firmware_upload_failed"]),
                    ("firmware_local_not_found", summary_map["firmware_local_not_found"]),
                ],
            ),
            (
                "inventory",
                [
                    ("inventory_ok", summary_map["inventory_ok"]),
                    (
                        "inventory_hostname_mismatch",
                        summary_map["inventory_hostname_mismatch"],
                    ),
                    (
                        "inventory_hostname_partial_match",
                        summary_map["inventory_hostname_partial_match"],
                    ),
                    (
                        "inventory_hostname_incomplete",
                        summary_map["inventory_hostname_incomplete"],
                    ),
                    ("inventory_not_found", summary_map["inventory_not_found"]),
                    ("inventory_duplicate", summary_map["inventory_duplicate"]),
                    ("severity_info", summary_map["severity_info"]),
                    ("severity_warning", summary_map["severity_warning"]),
                    ("severity_error", summary_map["severity_error"]),
                    ("match_ip", summary_map["match_ip"]),
                    ("match_hostname", summary_map["match_hostname"]),
                    ("match_partial_hostname", summary_map["match_partial_hostname"]),
                    ("match_not_found", summary_map["match_not_found"]),
                ],
            ),
        ]

    def _build_issue_rows(self, results: list[AuditResult]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []

        for result in results:
            has_audit_issue = not (
                result.status.startswith(AuditStatus.SSH_OK.value)
                or result.status.startswith(AuditStatus.FALLBACK_OK.value)
            )
            has_inventory_issue = (result.inventory_severity or "").upper() in {
                "WARNING",
                "ERROR",
            }
            has_firmware_issue = bool(result.firmware_error)

            if not (has_audit_issue or has_inventory_issue or has_firmware_issue):
                continue

            rows.append({header: getattr(result, header, "") for header in self.ISSUE_HEADERS})

        return rows

    @staticmethod
    def _build_rows(
        results: list[AuditResult],
        headers: list[str],
    ) -> list[dict[str, Any]]:
        return [
            {header: getattr(result, header, "") for header in headers}
            for result in results
        ]

    def _build_topology_rows(self, results: list[AuditResult]) -> list[dict[str, Any]]:
        rows = []

        for result in results:
            if not any(
                [
                    result.uplink_interface,
                    result.uplink_mac,
                    result.neighbor_identity,
                    result.neighbor_address,
                    result.neighbor_interface,
                    result.neighbor_mac,
                ]
            ):
                continue

            rows.append(
                {header: getattr(result, header, "") for header in self.TOPOLOGY_HEADERS}
            )

        return rows

    def _build_phpipam_mismatch_rows(
        self,
        results: list[AuditResult],
    ) -> list[dict[str, Any]]:
        rows = []

        for result in results:
            if (result.inventory_status or "OK") == "OK":
                continue

            rows.append(
                {
                    header: getattr(result, header, "")
                    for header in self.PHPIPAM_MISMATCH_HEADERS
                }
            )

        return rows

    def _populate_table_sheet(
        self,
        ws: Worksheet,
        headers: list[str],
        rows: list[dict[str, Any]],
        *,
        status_key: str | None = None,
        severity_key: str | None = None,
    ) -> None:
        ws.append(headers)
        self._style_header_row(ws, 1)

        for row in rows:
            ws.append([row.get(header, "") for header in headers])

            if severity_key:
                self._apply_issue_styles(ws, ws.max_row, row)
            elif status_key:
                self.apply_styles(ws, ws.max_row, str(row.get(status_key, "")))

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        self.autosize_worksheet(ws)

    def _resolve_json_path(self) -> str:
        output_json = getattr(self.config, "output_json", "")
        if output_json:
            return output_json

        xlsx_path = Path(self.config.output_xlsx)
        return str(xlsx_path.with_suffix(".json"))

    def export_json(self, results: list[AuditResult]) -> None:
        summary_rows = self._build_summary_rows(results)
        summary_dict = {key: value for key, value in summary_rows}
        topology_rows = self._build_topology_rows(results)
        mismatch_rows = self._build_phpipam_mismatch_rows(results)

        payload = {
            "summary": summary_dict,
            "report": {
                "generated_at_utc": datetime.now(timezone.utc).isoformat(),
                "inventory_file": self.config.inventory_file,
                "output_xlsx": self.config.output_xlsx,
                "issues": self._build_issue_rows(results),
                "topology": topology_rows,
                "phpipam_mismatches": mismatch_rows,
            },
            "results": [item.to_dict() for item in results],
        }

        output_json = self._resolve_json_path()

        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def export_xlsx(self, results: list[AuditResult]) -> None:
        wb = Workbook()

        ws = wb.active
        ws.title = "mikrotik_inventory"
        self._populate_table_sheet(
            ws,
            self.INVENTORY_HEADERS,
            self._build_rows(results, self.INVENTORY_HEADERS),
            status_key="status",
        )

        summary = wb.create_sheet("summary")
        summary.append(["metric", "value"])
        self._style_header_row(summary, 1)

        for section_name, rows in self._build_report_sections(results):
            summary.append([section_name.upper(), ""])
            self._style_section_row(summary, summary.max_row)

            for row in rows:
                summary.append(row)

            summary.append(["", ""])

        summary.freeze_panes = "A2"
        self.autosize_worksheet(summary)

        issues = wb.create_sheet("issues")
        self._populate_table_sheet(
            issues,
            self.ISSUE_HEADERS,
            self._build_issue_rows(results),
            severity_key="inventory_severity",
        )

        mismatches = wb.create_sheet("phpipam_mismatches")
        self._populate_table_sheet(
            mismatches,
            self.PHPIPAM_MISMATCH_HEADERS,
            self._build_phpipam_mismatch_rows(results),
            severity_key="inventory_severity",
        )

        topology = wb.create_sheet("topology")
        self._populate_table_sheet(
            topology,
            self.TOPOLOGY_HEADERS,
            self._build_topology_rows(results),
        )

        raw = wb.create_sheet("raw_inventory")
        self._populate_table_sheet(
            raw,
            self.RAW_HEADERS,
            self._build_rows(results, self.RAW_HEADERS),
            status_key="status",
        )

        wb.save(self.config.output_xlsx)

    def export(self, results: list[AuditResult]) -> None:
        self.export_xlsx(results)
        self.export_json(results)
