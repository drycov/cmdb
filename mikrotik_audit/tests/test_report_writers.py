"""Test cases for report writers behavior."""

from __future__ import annotations

import json

from openpyxl import load_workbook

from mikrotik_audit.report.writers.excel import ExcelWriter
from mikrotik_audit.report.writers.json import JsonWriter


def test_excel_writer_preserves_existing_sections_and_appends_rows(tmp_path) -> None:
    """Test that test excel writer preserves existing sections and appends rows."""
    report_path = tmp_path / "report.xlsx"

    first = ExcelWriter(str(report_path))
    first.open()
    first.begin_section("analyzer_summary", ["identity", "decision"])
    first.write_row("analyzer_summary", {"identity": "r1", "decision": "keep"})
    first.close_section("analyzer_summary")
    first.close()

    second = ExcelWriter(str(report_path))
    second.open()
    second.begin_section("topology_summary", ["total_devices"])
    second.write_row("topology_summary", {"total_devices": 3})
    second.close_section("topology_summary")
    second.close()

    third = ExcelWriter(str(report_path))
    third.open()
    third.begin_section("analyzer_summary", ["identity", "decision", "risks"])
    third.write_row(
        "analyzer_summary",
        {"identity": "r2", "decision": "append", "risks": "uplink"},
    )
    third.close_section("analyzer_summary")
    third.close()

    workbook = load_workbook(report_path)
    analyzer_sheet = workbook["analyzer_summary"]
    topology_sheet = workbook["topology_summary"]

    assert workbook.sheetnames == ["analyzer_summary", "topology_summary"]
    assert [cell.value for cell in analyzer_sheet[1]] == ["identity", "decision", "risks"]
    assert [cell.value for cell in analyzer_sheet[2]] == ["r1", "keep", None]
    assert [cell.value for cell in analyzer_sheet[3]] == ["r2", "append", "uplink"]
    assert [cell.value for cell in topology_sheet[1]] == ["total_devices"]
    assert [cell.value for cell in topology_sheet[2]] == [3]


def test_json_writer_appends_without_erasing_existing_rows(tmp_path) -> None:
    """Test that test json writer appends without erasing existing rows."""
    report_path = tmp_path / "report.ndjson"

    first = JsonWriter(str(report_path))
    first.open()
    first.begin_section("analyzer_summary", ["identity"])
    first.write_row("analyzer_summary", {"identity": "r1"})
    first.close_section("analyzer_summary")
    first.close()

    second = JsonWriter(str(report_path))
    second.open()
    second.begin_section("topology_summary", ["total_devices"])
    second.write_row("topology_summary", {"total_devices": 3})
    second.close_section("topology_summary")
    second.close()

    lines = report_path.read_text(encoding="utf-8").splitlines()

    assert [json.loads(line) for line in lines] == [
        {"section": "analyzer_summary", "identity": "r1"},
        {"section": "topology_summary", "total_devices": 3},
    ]


def test_excel_writer_recovers_from_corrupt_existing_workbook(tmp_path) -> None:
    """Test that test excel writer recovers from corrupt existing workbook."""
    report_path = tmp_path / "report.xlsx"
    report_path.write_text("not a real xlsx", encoding="utf-8")

    writer = ExcelWriter(str(report_path))
    writer.open()
    writer.begin_section("analyzer_summary", ["identity", "decision"])
    writer.write_row("analyzer_summary", {"identity": "r1", "decision": "keep"})
    writer.close_section("analyzer_summary")
    writer.close()

    backups = sorted(tmp_path.glob("report.corrupt-*.xlsx"))
    assert len(backups) == 1
    assert backups[0].read_text(encoding="utf-8") == "not a real xlsx"

    workbook = load_workbook(report_path)
    analyzer_sheet = workbook["analyzer_summary"]
    assert [cell.value for cell in analyzer_sheet[1]] == ["identity", "decision"]
    assert [cell.value for cell in analyzer_sheet[2]] == ["r1", "keep"]
