from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from config import AppConfig
from models import AuditResult

from .common import (
    INVENTORY_HEADERS,
    ISSUE_HEADERS,
    PHPIPAM_MISMATCH_HEADERS,
    TOPOLOGY_HEADERS,
    VLAN_HEADERS,
    build_issue_rows,
    build_phpipam_mismatch_rows,
    build_report_sections,
    build_summary_rows,
    build_topology_rows,
    build_vlan_rows,
    rows_by_headers,
)


class ExcelExporter:
    HEADER_FILL = PatternFill("solid", fgColor="1F1F1F")
    SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def _style_header(self, ws: Worksheet):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.HEADER_FILL

    def _populate(self, ws: Worksheet, headers, rows):
        ws.append(headers)
        self._style_header(ws)

        for row in rows:
            ws.append([row.get(h, "") for h in headers])

    def export_xlsx(self, results: list[AuditResult]) -> None:
        wb = Workbook()

        # inventory
        ws = wb.active
        ws.title = "mikrotik_inventory"
        self._populate(ws, INVENTORY_HEADERS, rows_by_headers(results, INVENTORY_HEADERS))

        # summary
        summary = wb.create_sheet("summary")
        summary.append(["metric", "value"])

        for section, rows in build_report_sections(
            results,
            inventory_file=self.config.inventory_file,
            output_xlsx=self.config.output_xlsx,
        ):
            summary.append([section.upper(), ""])
            for r in rows:
                summary.append(r)

        # issues
        issues = wb.create_sheet("issues")
        self._populate(issues, ISSUE_HEADERS, build_issue_rows(results))

        # mismatches
        mismatches = wb.create_sheet("phpipam_mismatches")
        self._populate(mismatches, PHPIPAM_MISMATCH_HEADERS, build_phpipam_mismatch_rows(results))

        # topology
        topology = wb.create_sheet("topology")
        self._populate(topology, TOPOLOGY_HEADERS, build_topology_rows(results))

        # vlans
        vlans = wb.create_sheet("vlans")
        self._populate(vlans, VLAN_HEADERS, build_vlan_rows(results))

        wb.save(self.config.output_xlsx)

    def export_json(self, results: list[AuditResult]) -> None:
        payload = {
            "summary": dict(build_summary_rows(results)),
            "report": {
                "generated_at_utc": datetime.now(timezone.utc).isoformat(),
                "inventory_file": self.config.inventory_file,
                "output_xlsx": self.config.output_xlsx,
                "issues": build_issue_rows(results),
                "topology": build_topology_rows(results),
                "phpipam_mismatches": build_phpipam_mismatch_rows(results),
                "vlans": build_vlan_rows(results),
            },
            "results": [r.to_dict() for r in results],
        }

        path = Path(self.config.output_xlsx).with_suffix(".json")

        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)

    def export(self, results: list[AuditResult]) -> None:
        self.export_xlsx(results)
        self.export_json(results)