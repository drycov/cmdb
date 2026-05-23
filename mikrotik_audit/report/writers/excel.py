"""Implementation details for report writers excel."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from mikrotik_audit.report.writers.tabular import merge_headers, project_row


class ExcelWriter:
    """Write excelwriter output."""
    def __init__(self, path: str) -> None:
        self.path = path
        self.wb = None
        self.sheets = {}
        self.headers = {}
        self._created_new = False
        self.logger = logging.getLogger(__name__)

    def open(self) -> None:
        workbook_path = Path(self.path)
        if workbook_path.exists():
            try:
                self.wb = load_workbook(workbook_path)
                self._created_new = False
                return
            except (BadZipFile, InvalidFileException, KeyError, OSError) as exc:
                backup_path = self._backup_corrupt_workbook(workbook_path)
                self.logger.warning(
                    "Excel workbook '%s' is invalid; archived it to '%s' and creating a new workbook. error=%s",
                    workbook_path,
                    backup_path,
                    exc,
                )

        self.wb = Workbook()
        self._created_new = True

    def begin_section(self, name: str, headers: list[str]) -> None:
        self._remove_default_sheet()

        if name in self.wb.sheetnames:
            ws = self.wb[name]
            merged_headers = self._ensure_headers(ws, headers)
        else:
            ws = self.wb.create_sheet(title=name)
            ws.append(headers)
            merged_headers = headers

        self.sheets[name] = ws
        self.headers[name] = merged_headers

    def write_row(self, section: str, row: dict[str, Any]) -> None:
        ws = self.sheets[section]
        headers = self.headers[section]

        ws.append(project_row(headers, row))

    def close_section(self, name: str) -> None:
        pass

    def close(self) -> None:
        if self.wb is None:
            return
        self.wb.save(self.path)

    def _remove_default_sheet(self) -> None:
        if not self._created_new:
            return

        if self.wb.sheetnames != ["Sheet"]:
            return

        sheet = self.wb["Sheet"]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
            self.wb.remove(sheet)

    def _ensure_headers(self, ws, headers: list[str]) -> list[str]:
        existing_headers = self._read_headers(ws)
        if not existing_headers:
            ws.append(headers)
            return headers

        merged_headers = merge_headers(existing_headers, headers)
        for index, header in enumerate(merged_headers, start=1):
            ws.cell(row=1, column=index, value=header)
        return merged_headers

    @staticmethod
    def _read_headers(ws) -> list[str]:
        if ws.max_row < 1:
            return []

        values = [cell.value for cell in ws[1]]
        if not any(value is not None and value != "" for value in values):
            return []
        return ["" if value is None else str(value) for value in values]

    @staticmethod
    def _backup_corrupt_workbook(workbook_path: Path) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = workbook_path.with_name(
            f"{workbook_path.stem}.corrupt-{timestamp}{workbook_path.suffix}"
        )
        workbook_path.replace(backup_path)
        return backup_path
