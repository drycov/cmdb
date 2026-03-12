from __future__ import annotations

import json
from pathlib import Path
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import AppConfig
from constants.error_codes import FirmwareErrorCode
from constants.statuses import AuditStatus
from models import AuditResult


class ExcelExporter:
    def __init__(self, config: AppConfig) -> None:
        self.config = config

    @staticmethod
    def autosize_worksheet(ws) -> None:
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)

            for cell in col_cells:
                try:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    @staticmethod
    def apply_styles(ws, row_idx: int, status: str) -> None:
        fills = {
            AuditStatus.SSH_OK.value: PatternFill("solid", fgColor="C6EFCE"),
            AuditStatus.FALLBACK_OK.value: PatternFill("solid", fgColor="FFF2CC"),
            AuditStatus.AUTH_FAILED.value: PatternFill("solid", fgColor="F4CCCC"),
            AuditStatus.OFFLINE.value: PatternFill("solid", fgColor="D9D9D9"),
            AuditStatus.SSH_CLOSED.value: PatternFill("solid", fgColor="FCE5CD"),
        }

        fill = None
        if status.startswith(AuditStatus.SSH_OK.value):
            fill = fills[AuditStatus.SSH_OK.value]
        elif status.startswith(AuditStatus.FALLBACK_OK.value):
            fill = fills[AuditStatus.FALLBACK_OK.value]
        elif status == AuditStatus.AUTH_FAILED.value:
            fill = fills[AuditStatus.AUTH_FAILED.value]
        elif status == AuditStatus.OFFLINE.value:
            fill = fills[AuditStatus.OFFLINE.value]
        elif status == AuditStatus.SSH_CLOSED.value:
            fill = fills[AuditStatus.SSH_CLOSED.value]

        if fill:
            for cell in ws[row_idx]:
                cell.fill = fill

    def _build_summary_rows(self, results: List[AuditResult]) -> list[tuple[str, Any]]:
        return [
            ("total_hosts", len(results)),
            ("alive", sum(1 for r in results if r.ping)),
            (
                "ssh_ok",
                sum(1 for r in results if r.status.startswith(AuditStatus.SSH_OK.value)),
            ),
            (
                "fallback_ok",
                sum(1 for r in results if r.status.startswith(AuditStatus.FALLBACK_OK.value)),
            ),
            (
                "auth_failed",
                sum(1 for r in results if r.status == AuditStatus.AUTH_FAILED.value),
            ),
            (
                "ssh_closed",
                sum(1 for r in results if r.status == AuditStatus.SSH_CLOSED.value),
            ),
            (
                "offline",
                sum(1 for r in results if r.status == AuditStatus.OFFLINE.value),
            ),
            ("radius_added", sum(1 for r in results if r.radius_added)),
            ("radius_recreated", sum(1 for r in results if r.radius_recreated)),
            ("radius_present_after", sum(1 for r in results if r.radius_present_after)),
            ("aaa_enabled", sum(1 for r in results if r.aaa_enabled)),
            ("aaa_present_after", sum(1 for r in results if r.aaa_present_after)),
            ("firmware_upload_needed", sum(1 for r in results if r.firmware_upload_needed)),
            ("firmware_uploaded", sum(1 for r in results if r.firmware_uploaded)),
            (
                "firmware_already_present",
                sum(1 for r in results if r.firmware_already_present),
            ),
            ("firmware_reboot_sent", sum(1 for r in results if r.firmware_reboot_sent)),
            (
                "firmware_same_version",
                sum(
                    1
                    for r in results
                    if r.firmware_error == FirmwareErrorCode.SAME_VERSION.value
                ),
            ),
            (
                "firmware_upload_failed",
                sum(
                    1
                    for r in results
                    if r.firmware_error == FirmwareErrorCode.UPLOAD_FAILED.value
                ),
            ),
            (
                "firmware_local_not_found",
                sum(
                    1
                    for r in results
                    if r.firmware_error == FirmwareErrorCode.LOCAL_FIRMWARE_NOT_FOUND.value
                ),
            ),
        ]

    def _resolve_json_path(self) -> str:
        output_json = getattr(self.config, "output_json", "")
        if output_json:
            return output_json

        xlsx_path = Path(self.config.output_xlsx)
        return str(xlsx_path.with_suffix(".json"))

    def export_json(self, results: List[AuditResult]) -> None:
        summary_rows = self._build_summary_rows(results)
        summary_dict = {key: value for key, value in summary_rows}

        payload = {
            "summary": summary_dict,
            "results": [item.to_dict() for item in results],
        }

        output_json = self._resolve_json_path()

        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def export_xlsx(self, results: List[AuditResult]) -> None:
        wb = Workbook()

        ws = wb.active
        ws.title = "mikrotik_inventory"

        headers = AuditResult.EXPORT_HEADERS
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for item in results:
            row_dict = item.to_dict()
            ws.append([row_dict.get(header, "") for header in headers])
            self.apply_styles(ws, ws.max_row, str(item.status))

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        self.autosize_worksheet(ws)

        summary = wb.create_sheet("summary")
        summary.append(["metric", "value"])

        for cell in summary[1]:
            cell.font = Font(bold=True)

        summary_rows = self._build_summary_rows(results)
        for row in summary_rows:
            summary.append(row)

        self.autosize_worksheet(summary)
        wb.save(self.config.output_xlsx)

    def export(self, results: List[AuditResult]) -> None:
        self.export_xlsx(results)
        self.export_json(results)