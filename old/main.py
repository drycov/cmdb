import asyncio
import ipaddress
import logging
import os
import socket
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Dict, List, Optional

import paramiko
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pythonping import ping
from tqdm import tqdm


# =========================
# CONFIG
# =========================

load_dotenv()

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
LOG_DIR = os.getenv("LOG_DIR", "logs")
LOG_FILE = os.getenv("LOG_FILE", "mikrotik_audit.log")
ERROR_LOG_FILE = os.getenv("ERROR_LOG_FILE", "mikrotik_audit.error.log")

USERNAME = os.getenv("MIKROTIK_USERNAME", "")
PASSWORD = os.getenv("MIKROTIK_PASSWORD", "")

SSH_PORT = int(os.getenv("MIKROTIK_SSH_PORT", 22))
TIMEOUT = int(os.getenv("MIKROTIK_TIMEOUT", 2))
WORKERS = int(os.getenv("MIKROTIK_WORKERS", 100))

FALLBACK_USERNAME = os.getenv("FALLBACK_USERNAME", "satcoadm")
FALLBACK_PASSWORD = os.getenv("FALLBACK_PASSWORD", "password")

FIRMWARE_USERNAME = FALLBACK_USERNAME
FIRMWARE_PASSWORD = FALLBACK_PASSWORD

RADIUS_ADDR = os.getenv("RADIUS_ADDR", "10.216.40.3")
RADIUS_SECRET = os.getenv("RADIUS_SECRET", "secret")
RADIUS_SERVICE = os.getenv("RADIUS_SERVICE", "login")

OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "mikrotik_inventory.xlsx")

FIRMWARE_DIR = os.getenv("FIRMWARE_DIR", "firmware")
AUTO_UPLOAD_MMIPS = os.getenv("AUTO_UPLOAD_MMIPS", "false").lower() == "true"
AUTO_REBOOT_AFTER_UPLOAD = os.getenv("AUTO_REBOOT_AFTER_UPLOAD", "false").lower() == "true"
ONLY_IF_VERSION_DIFF = os.getenv("ONLY_IF_VERSION_DIFF", "true").lower() == "true"

TEST_MODE = os.getenv("TEST_MODE", "false").lower() == "true"
TEST_LIMIT = int(os.getenv("TEST_LIMIT", 2))
TEST_IPS_RAW = os.getenv("TEST_IPS", "").strip()

SUBNETS_FILE = os.getenv("SUBNETS_FILE", "subnets.txt")

executor = ThreadPoolExecutor(max_workers=WORKERS)

logging.getLogger("paramiko.transport").setLevel(logging.CRITICAL)
logging.getLogger("paramiko").setLevel(logging.CRITICAL)


# =========================
# LOGGING
# =========================

def setup_logging() -> logging.Logger:
    os.makedirs(LOG_DIR, exist_ok=True)

    logger_instance = logging.getLogger("mikrotik_audit")
    logger_instance.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    logger_instance.handlers.clear()

    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(threadName)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(
        os.path.join(LOG_DIR, LOG_FILE),
        encoding="utf-8",
    )
    file_handler.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    file_handler.setFormatter(formatter)

    error_file_handler = logging.FileHandler(
        os.path.join(LOG_DIR, ERROR_LOG_FILE),
        encoding="utf-8",
    )
    error_file_handler.setLevel(logging.ERROR)
    error_file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger_instance.addHandler(file_handler)
    logger_instance.addHandler(error_file_handler)
    logger_instance.addHandler(console_handler)
    logger_instance.propagate = False

    return logger_instance


logger = setup_logging()


# =========================
# HELPERS
# =========================

def network_of_ip(ip: str) -> str:
    parts = ip.split(".")
    return f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"


def generate_ips() -> List[str]:
    ips: List[str] = []

    with open(SUBNETS_FILE, encoding="utf-8") as f:
        for raw in f:
            subnet = raw.strip()
            if not subnet or subnet.startswith("#"):
                continue

            net = ipaddress.ip_network(subnet, strict=False)
            for ip in net.hosts():
                ips.append(str(ip))

    return ips


def get_target_ips() -> List[str]:
    ips = generate_ips()

    if not TEST_MODE:
        return ips

    if TEST_IPS_RAW:
        selected = [ip.strip() for ip in TEST_IPS_RAW.split(",") if ip.strip()]
        return selected

    return ips[:TEST_LIMIT]


def ping_host(ip: str) -> bool:
    try:
        result = ping(ip, count=1, timeout=1)
        return result.success()
    except Exception as exc:
        logger.debug("Ping exception ip=%s error=%s", ip, exc)
        return False


def check_ssh_port(ip: str) -> bool:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(TIMEOUT)
            sock.connect((ip, SSH_PORT))
            return True
    except Exception as exc:
        logger.debug("SSH port check failed ip=%s port=%s error=%s", ip, SSH_PORT, exc)
        return False


def ssh_exec(ip: str, username: str, password: str, command: str) -> Optional[str]:
    client = None
    try:
        logger.debug("SSH exec start ip=%s user=%s cmd=%s", ip, username, command)

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        client.connect(
            ip,
            port=SSH_PORT,
            username=username,
            password=password,
            timeout=TIMEOUT,
            banner_timeout=TIMEOUT,
            auth_timeout=TIMEOUT,
            look_for_keys=False,
            allow_agent=False,
        )

        _, stdout, stderr = client.exec_command(command, timeout=TIMEOUT + 5)

        output = stdout.read().decode(errors="ignore")
        err = stderr.read().decode(errors="ignore")

        if err and not output:
            logger.warning(
                "SSH command returned stderr only ip=%s user=%s cmd=%s err=%s",
                ip,
                username,
                command,
                err.strip(),
            )
            return None

        logger.debug("SSH exec success ip=%s user=%s cmd=%s", ip, username, command)
        return output

    except Exception as exc:
        logger.debug(
            "SSH exec exception ip=%s user=%s cmd=%s error=%s",
            ip,
            username,
            command,
            exc,
        )
        return None

    finally:
        if client:
            try:
                client.close()
            except Exception:
                pass


def parse_colon_output(output: str) -> Dict[str, str]:
    data: Dict[str, str] = {}

    for raw_line in output.splitlines():
        line = raw_line.strip()
        if not line or ":" not in line:
            continue

        key, value = line.split(":", 1)
        key = key.strip().lower().replace(" ", "_")
        value = value.strip()
        data[key] = value

    return data


def normalize_version(version: str) -> str:
    return version.strip()


def find_firmware_file(architecture: str) -> Optional[Path]:
    fw_dir = Path(FIRMWARE_DIR)
    if not fw_dir.exists() or not fw_dir.is_dir():
        return None

    patterns = [
        f"routeros-{architecture}-*.npk",
        f"*{architecture}*.npk",
    ]

    for pattern in patterns:
        candidates = sorted(fw_dir.glob(pattern))
        if candidates:
            return candidates[0]

    return None


def extract_version_from_filename(filename: str, architecture: str) -> str:
    patterns = [
        (f"routeros-{architecture}-", ".npk"),
        ("routeros-", f"-{architecture}.npk"),
    ]

    name = filename.strip()
    for prefix, suffix in patterns:
        if name.startswith(prefix) and name.endswith(suffix):
            return name[len(prefix): len(name) - len(suffix)]

    return ""


def upload_file_sftp(
    ip: str,
    username: str,
    password: str,
    local_path: Path,
    remote_name: Optional[str] = None,
) -> bool:
    client = None
    sftp = None

    try:
        logger.info(
            "SFTP upload start ip=%s user=%s file=%s",
            ip,
            username,
            local_path.name,
        )

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        client.connect(
            ip,
            port=SSH_PORT,
            username=username,
            password=password,
            timeout=TIMEOUT,
            banner_timeout=TIMEOUT,
            auth_timeout=TIMEOUT,
            look_for_keys=False,
            allow_agent=False,
        )

        sftp = client.open_sftp()
        sftp.put(str(local_path), remote_name or local_path.name)

        logger.info(
            "SFTP upload success ip=%s user=%s file=%s",
            ip,
            username,
            local_path.name,
        )
        return True

    except Exception as exc:
        logger.error(
            "SFTP upload failed ip=%s user=%s file=%s error=%s",
            ip,
            username,
            local_path.name,
            exc,
        )
        return False

    finally:
        if sftp:
            try:
                sftp.close()
            except Exception:
                pass
        if client:
            try:
                client.close()
            except Exception:
                pass


def remote_file_exists(ip: str, username: str, password: str, filename: str) -> bool:
    out = ssh_exec(ip, username, password, f'/file print where name="{filename}"')
    if not out:
        return False
    return filename.lower() in out.lower()


# =========================
# FIRMWARE
# =========================

def ensure_firmware_uploaded(
    ip: str,
    username: str,
    password: str,
    architecture: str,
    current_version: str,
) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "firmware_candidate": "",
        "firmware_target_version": "",
        "firmware_upload_needed": False,
        "firmware_uploaded": False,
        "firmware_already_present": False,
        "firmware_reboot_sent": False,
        "firmware_error": "",
    }

    logger.info(
        "Firmware check ip=%s arch=%s current_version=%s",
        ip,
        architecture,
        current_version,
    )

    if architecture != "mmips":
        result["firmware_error"] = "skip_non_mmips"
        logger.info("Firmware skipped ip=%s reason=non_mmips arch=%s", ip, architecture)
        return result

    fw = find_firmware_file("mmips")
    if not fw:
        result["firmware_error"] = "local_firmware_not_found"
        logger.error(
            "Firmware file not found ip=%s arch=%s dir=%s",
            ip,
            architecture,
            FIRMWARE_DIR,
        )
        return result

    result["firmware_candidate"] = fw.name
    result["firmware_target_version"] = extract_version_from_filename(fw.name, "mmips")

    normalized_current = normalize_version(current_version)
    normalized_target = normalize_version(result["firmware_target_version"])

    if ONLY_IF_VERSION_DIFF and normalized_current and normalized_target:
        if normalized_current == normalized_target:
            result["firmware_error"] = "same_version"
            logger.info(
                "Firmware skipped ip=%s reason=same_version current=%s target=%s",
                ip,
                normalized_current,
                normalized_target,
            )
            return result

    result["firmware_upload_needed"] = True

    if remote_file_exists(ip, username, password, fw.name):
        result["firmware_already_present"] = True
        logger.info("Firmware already present ip=%s file=%s", ip, fw.name)
    else:
        uploaded = upload_file_sftp(ip, username, password, fw)
        if not uploaded:
            result["firmware_error"] = "upload_failed"
            return result
        result["firmware_uploaded"] = True

    if AUTO_REBOOT_AFTER_UPLOAD:
        logger.warning("Sending reboot ip=%s", ip)
        reboot_ok = ssh_exec(ip, username, password, "/system reboot") is not None
        result["firmware_reboot_sent"] = reboot_ok
        if not reboot_ok:
            result["firmware_error"] = "reboot_command_failed"
            logger.error("Reboot command failed ip=%s", ip)

    return result


# =========================
# RADIUS REMEDIATION
# =========================

def ensure_radius(ip: str) -> Dict[str, bool]:
    result = {
        "radius_added": False,
        "radius_recreated": False,
        "radius_present_after": False,
        "aaa_enabled": False,
        "aaa_present_after": False,
    }

    logger.info("RADIUS remediation start ip=%s", ip)

    count_cmd = f'/radius print count-only where service={RADIUS_SERVICE} and address="{RADIUS_ADDR}"'
    radius_count_raw = ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, count_cmd)

    radius_count = 0
    if radius_count_raw:
        lines = [line.strip() for line in radius_count_raw.splitlines() if line.strip()]
        if lines:
            try:
                radius_count = int(lines[-1])
            except ValueError:
                radius_count = 0

    logger.info("RADIUS current count ip=%s count=%s", ip, radius_count)

    if radius_count == 0:
        add_cmd = (
            f"/radius add service={RADIUS_SERVICE} "
            f"address={RADIUS_ADDR} secret={RADIUS_SECRET}"
        )
        ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, add_cmd)
        result["radius_added"] = True
        logger.info("RADIUS entry added ip=%s", ip)

    elif radius_count > 1:
        remove_cmd = f'/radius remove [find where service={RADIUS_SERVICE} and address="{RADIUS_ADDR}"]'
        ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, remove_cmd)

        add_cmd = (
            f"/radius add service={RADIUS_SERVICE} "
            f"address={RADIUS_ADDR} secret={RADIUS_SECRET}"
        )
        ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, add_cmd)
        result["radius_recreated"] = True
        logger.info("RADIUS entries recreated ip=%s", ip)

    radius_verify_raw = ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, count_cmd)

    radius_verify_count = 0
    if radius_verify_raw:
        lines = [line.strip() for line in radius_verify_raw.splitlines() if line.strip()]
        if lines:
            try:
                radius_verify_count = int(lines[-1])
            except ValueError:
                radius_verify_count = 0

    result["radius_present_after"] = radius_verify_count == 1

    aaa_before = ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, "/user aaa print")
    aaa_enabled_before = bool(aaa_before and "use-radius: yes" in aaa_before.lower())

    if not aaa_enabled_before:
        ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, "/user aaa set use-radius=yes")
        result["aaa_enabled"] = True
        logger.info("AAA use-radius enabled ip=%s", ip)

    aaa_after = ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, "/user aaa print")
    result["aaa_present_after"] = bool(
        aaa_after and "use-radius: yes" in aaa_after.lower()
    )

    logger.info(
        "RADIUS remediation finished ip=%s radius_present_after=%s aaa_present_after=%s",
        ip,
        result["radius_present_after"],
        result["aaa_present_after"],
    )

    return result


# =========================
# DATA COLLECTION
# =========================

def collect_router_data(
    ip: str,
    username: str,
    password: str,
) -> Optional[Dict[str, Any]]:
    resource_out = ssh_exec(ip, username, password, "/system resource print")
    if resource_out is None:
        return None

    identity_out = ssh_exec(ip, username, password, "/system identity print")
    routerboard_out = ssh_exec(ip, username, password, "/system routerboard print")
    license_out = ssh_exec(ip, username, password, "/system license print")
    interfaces_out = ssh_exec(ip, username, password, "/interface print count-only")

    resource = parse_colon_output(resource_out)
    identity = parse_colon_output(identity_out or "")
    routerboard = parse_colon_output(routerboard_out or "")
    license_data = parse_colon_output(license_out or "")

    interface_count = ""
    if interfaces_out:
        interface_count = interfaces_out.strip().splitlines()[-1].strip()

    return {
        "identity": identity.get("name", ""),
        "version": resource.get("version", ""),
        "uptime": resource.get("uptime", ""),
        "cpu_load": resource.get("cpu-load", ""),
        "board_name": resource.get("board-name", ""),
        "platform": resource.get("platform", ""),
        "architecture": resource.get("architecture-name", ""),
        "total_memory": resource.get("total-memory", ""),
        "free_memory": resource.get("free-memory", ""),
        "total_hdd": resource.get("total-hdd-space", ""),
        "free_hdd": resource.get("free-hdd-space", ""),
        "license": license_data.get("software-id", "") or license_data.get("level", ""),
        "current_firmware": routerboard.get("current-firmware", ""),
        "upgrade_firmware": routerboard.get("upgrade-firmware", ""),
        "interface_count": interface_count,
    }


# =========================
# AUDIT
# =========================

def audit_device(ip: str) -> Dict[str, Any]:
    logger.info("Audit started ip=%s", ip)

    result: Dict[str, Any] = {
        "ip": ip,
        "subnet": network_of_ip(ip),
        "identity": "",
        "ping": False,
        "ssh_port": False,
        "version": "",
        "uptime": "",
        "cpu_load": "",
        "board_name": "",
        "platform": "",
        "architecture": "",
        "total_memory": "",
        "free_memory": "",
        "total_hdd": "",
        "free_hdd": "",
        "license": "",
        "current_firmware": "",
        "upgrade_firmware": "",
        "interface_count": "",
        "auth_method": "",
        "radius_added": False,
        "radius_recreated": False,
        "radius_present_after": False,
        "aaa_enabled": False,
        "aaa_present_after": False,
        "firmware_candidate": "",
        "firmware_target_version": "",
        "firmware_upload_needed": False,
        "firmware_uploaded": False,
        "firmware_already_present": False,
        "firmware_reboot_sent": False,
        "firmware_error": "",
        "status": "",
    }

    if not ping_host(ip):
        result["status"] = "offline"
        logger.info("Audit finished ip=%s status=%s", ip, result["status"])
        return result

    result["ping"] = True

    if not check_ssh_port(ip):
        result["status"] = "ssh_closed"
        logger.info("Audit finished ip=%s status=%s", ip, result["status"])
        return result

    result["ssh_port"] = True

    primary_data = collect_router_data(ip, USERNAME, PASSWORD)
    if primary_data:
        result.update(primary_data)
        result["auth_method"] = "primary"

        logger.info(
            "Primary auth success ip=%s identity=%s version=%s arch=%s",
            ip,
            result["identity"],
            result["version"],
            result["architecture"],
        )

        if AUTO_UPLOAD_MMIPS:
            fw_result = ensure_firmware_uploaded(
                ip=ip,
                username=FIRMWARE_USERNAME,
                password=FIRMWARE_PASSWORD,
                architecture=result.get("architecture", ""),
                current_version=result.get("version", ""),
            )
            result.update(fw_result)

        status_parts = ["ssh_ok"]

        if result["firmware_uploaded"]:
            status_parts.append("fw_uploaded")
        if result["firmware_already_present"]:
            status_parts.append("fw_already_present")
        if result["firmware_reboot_sent"]:
            status_parts.append("fw_reboot_sent")
        if result["firmware_error"] and not result["firmware_error"].startswith("skip_"):
            status_parts.append(result["firmware_error"])

        result["status"] = "_".join(status_parts)

        logger.info(
            "DEVICE_RESULT ip=%s status=%s auth=%s identity=%s version=%s arch=%s fw_uploaded=%s radius_added=%s aaa_enabled=%s",
            result["ip"],
            result["status"],
            result["auth_method"],
            result["identity"],
            result["version"],
            result["architecture"],
            result["firmware_uploaded"],
            result["radius_added"],
            result["aaa_enabled"],
        )
        return result

    logger.info("Primary auth failed, trying fallback ip=%s", ip)

    fallback_identity = ssh_exec(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD, "/system identity print")
    if fallback_identity is None:
        result["status"] = "auth_failed"
        logger.warning("Audit finished ip=%s status=%s", ip, result["status"])
        return result

    fallback_data = collect_router_data(ip, FALLBACK_USERNAME, FALLBACK_PASSWORD)
    if fallback_data:
        result.update(fallback_data)

    result["auth_method"] = "fallback"

    logger.info(
        "Fallback auth success ip=%s identity=%s version=%s arch=%s",
        ip,
        result["identity"],
        result["version"],
        result["architecture"],
    )

    if AUTO_UPLOAD_MMIPS:
        fw_result = ensure_firmware_uploaded(
            ip=ip,
            username=FIRMWARE_USERNAME,
            password=FIRMWARE_PASSWORD,
            architecture=result.get("architecture", ""),
            current_version=result.get("version", ""),
        )
        result.update(fw_result)

    fix_result = ensure_radius(ip)
    result["radius_added"] = fix_result["radius_added"]
    result["radius_recreated"] = fix_result["radius_recreated"]
    result["radius_present_after"] = fix_result["radius_present_after"]
    result["aaa_enabled"] = fix_result["aaa_enabled"]
    result["aaa_present_after"] = fix_result["aaa_present_after"]

    status_parts = ["fallback_ok"]

    if result["radius_added"]:
        status_parts.append("radius_added")
    if result["radius_recreated"]:
        status_parts.append("radius_recreated")
    if result["aaa_enabled"]:
        status_parts.append("aaa_enabled")
    if not result["radius_present_after"]:
        status_parts.append("radius_verify_failed")
    if not result["aaa_present_after"]:
        status_parts.append("aaa_verify_failed")
    if result["firmware_uploaded"]:
        status_parts.append("fw_uploaded")
    if result["firmware_already_present"]:
        status_parts.append("fw_already_present")
    if result["firmware_reboot_sent"]:
        status_parts.append("fw_reboot_sent")
    if result["firmware_error"] and not result["firmware_error"].startswith("skip_"):
        status_parts.append(result["firmware_error"])

    result["status"] = "_".join(status_parts)

    logger.info(
        "DEVICE_RESULT ip=%s status=%s auth=%s identity=%s version=%s arch=%s fw_uploaded=%s radius_added=%s aaa_enabled=%s",
        result["ip"],
        result["status"],
        result["auth_method"],
        result["identity"],
        result["version"],
        result["architecture"],
        result["firmware_uploaded"],
        result["radius_added"],
        result["aaa_enabled"],
    )

    return result


# =========================
# XLSX EXPORT
# =========================

def autosize_worksheet(ws) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def apply_styles(ws, row_idx: int, status: str) -> None:
    fills = {
        "ssh_ok": PatternFill("solid", fgColor="C6EFCE"),
        "fallback": PatternFill("solid", fgColor="FFF2CC"),
        "auth_failed": PatternFill("solid", fgColor="F4CCCC"),
        "offline": PatternFill("solid", fgColor="D9D9D9"),
        "ssh_closed": PatternFill("solid", fgColor="FCE5CD"),
    }

    fill = None
    if status.startswith("ssh_ok"):
        fill = fills["ssh_ok"]
    elif status.startswith("fallback_ok"):
        fill = fills["fallback"]
    elif status == "auth_failed":
        fill = fills["auth_failed"]
    elif status == "offline":
        fill = fills["offline"]
    elif status == "ssh_closed":
        fill = fills["ssh_closed"]

    if fill:
        for cell in ws[row_idx]:
            cell.fill = fill


def export_xlsx(results: List[Dict[str, Any]]) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "mikrotik_inventory"

    headers = [
        "ip",
        "subnet",
        "identity",
        "ping",
        "ssh_port",
        "auth_method",
        "version",
        "uptime",
        "cpu_load",
        "board_name",
        "platform",
        "architecture",
        "total_memory",
        "free_memory",
        "total_hdd",
        "free_hdd",
        "license",
        "current_firmware",
        "upgrade_firmware",
        "interface_count",
        "radius_added",
        "radius_recreated",
        "radius_present_after",
        "aaa_enabled",
        "aaa_present_after",
        "firmware_candidate",
        "firmware_target_version",
        "firmware_upload_needed",
        "firmware_uploaded",
        "firmware_already_present",
        "firmware_reboot_sent",
        "firmware_error",
        "status",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in results:
        ws.append(
            [
                item["ip"],
                item["subnet"],
                item["identity"],
                item["ping"],
                item["ssh_port"],
                item["auth_method"],
                item["version"],
                item["uptime"],
                item["cpu_load"],
                item["board_name"],
                item["platform"],
                item["architecture"],
                item["total_memory"],
                item["free_memory"],
                item["total_hdd"],
                item["free_hdd"],
                item["license"],
                item["current_firmware"],
                item["upgrade_firmware"],
                item["interface_count"],
                item["radius_added"],
                item["radius_recreated"],
                item["radius_present_after"],
                item["aaa_enabled"],
                item["aaa_present_after"],
                item["firmware_candidate"],
                item["firmware_target_version"],
                item["firmware_upload_needed"],
                item["firmware_uploaded"],
                item["firmware_already_present"],
                item["firmware_reboot_sent"],
                item["firmware_error"],
                item["status"],
            ]
        )
        apply_styles(ws, ws.max_row, item["status"])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])

    for cell in summary[1]:
        cell.font = Font(bold=True)

    total = len(results)
    alive = sum(1 for r in results if r["ping"])
    ssh_ok = sum(1 for r in results if str(r["status"]).startswith("ssh_ok"))
    fallback_ok = sum(1 for r in results if str(r["status"]).startswith("fallback_ok"))
    radius_added = sum(1 for r in results if r["radius_added"])
    radius_recreated = sum(1 for r in results if r["radius_recreated"])
    radius_present_after = sum(1 for r in results if r["radius_present_after"])
    aaa_enabled = sum(1 for r in results if r["aaa_enabled"])
    aaa_present_after = sum(1 for r in results if r["aaa_present_after"])
    auth_failed = sum(1 for r in results if r["status"] == "auth_failed")
    ssh_closed = sum(1 for r in results if r["status"] == "ssh_closed")
    offline = sum(1 for r in results if r["status"] == "offline")
    radius_verify_failed = sum(
        1 for r in results if "radius_verify_failed" in str(r["status"])
    )
    aaa_verify_failed = sum(
        1 for r in results if "aaa_verify_failed" in str(r["status"])
    )
    firmware_upload_needed = sum(1 for r in results if r["firmware_upload_needed"])
    firmware_uploaded = sum(1 for r in results if r["firmware_uploaded"])
    firmware_already_present = sum(1 for r in results if r["firmware_already_present"])
    firmware_reboot_sent = sum(1 for r in results if r["firmware_reboot_sent"])
    firmware_same_version = sum(1 for r in results if r["firmware_error"] == "same_version")
    firmware_upload_failed = sum(1 for r in results if r["firmware_error"] == "upload_failed")
    firmware_local_not_found = sum(
        1 for r in results if r["firmware_error"] == "local_firmware_not_found"
    )

    summary_rows = [
        ("total_hosts", total),
        ("alive", alive),
        ("ssh_ok", ssh_ok),
        ("fallback_ok", fallback_ok),
        ("radius_added", radius_added),
        ("radius_recreated", radius_recreated),
        ("radius_present_after", radius_present_after),
        ("aaa_enabled", aaa_enabled),
        ("aaa_present_after", aaa_present_after),
        ("auth_failed", auth_failed),
        ("ssh_closed", ssh_closed),
        ("offline", offline),
        ("radius_verify_failed", radius_verify_failed),
        ("aaa_verify_failed", aaa_verify_failed),
        ("firmware_upload_needed", firmware_upload_needed),
        ("firmware_uploaded", firmware_uploaded),
        ("firmware_already_present", firmware_already_present),
        ("firmware_reboot_sent", firmware_reboot_sent),
        ("firmware_same_version", firmware_same_version),
        ("firmware_upload_failed", firmware_upload_failed),
        ("firmware_local_not_found", firmware_local_not_found),
    ]

    for row in summary_rows:
        summary.append(row)

    autosize_worksheet(summary)
    wb.save(OUTPUT_XLSX)


# =========================
# RUNNER
# =========================

async def run_audit() -> None:
    logger.info("Run started")

    ips = get_target_ips()

    if not ips:
        logger.warning("No target IPs found")
        return

    if TEST_MODE and AUTO_UPLOAD_MMIPS and len(ips) > 2:
        logger.error("Refusing firmware upload in TEST_MODE for more than 2 devices")
        return

    logger.info("Total target IPs: %s", len(ips))
    for ip in ips:
        logger.info("Queue target ip=%s", ip)

    loop = asyncio.get_event_loop()
    tasks = [loop.run_in_executor(executor, audit_device, ip) for ip in ips]

    results: List[Dict[str, Any]] = []

    stats = {
        "alive": 0,
        "ssh_ok": 0,
        "fallback": 0,
        "fail": 0,
    }

    with tqdm(total=len(tasks), desc="Scanning MikroTik") as pbar:
        for future in asyncio.as_completed(tasks):
            result = await future
            results.append(result)

            if result["ping"]:
                stats["alive"] += 1

            if str(result["status"]).startswith("ssh_ok"):
                stats["ssh_ok"] += 1
            elif str(result["status"]).startswith("fallback_ok"):
                stats["fallback"] += 1
            elif result["status"] != "offline":
                stats["fail"] += 1

            pbar.set_postfix(stats)
            pbar.update(1)

    results.sort(key=lambda x: tuple(int(p) for p in x["ip"].split(".")))

    logger.info("Exporting XLSX rows=%s file=%s", len(results), OUTPUT_XLSX)
    export_xlsx(results)
    logger.info("Run finished successfully rows=%s file=%s", len(results), OUTPUT_XLSX)


if __name__ == "__main__":
    try:
        logger.info("Application started")
        asyncio.run(run_audit())
        logger.info("Application finished")
    except Exception as exc:
        logger.exception("Application crashed error=%s", exc)
        raise